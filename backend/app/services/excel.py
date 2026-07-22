"""Excel (.xlsx) and CSV export."""

from __future__ import annotations

import csv
import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.logging_config import get_logger

log = get_logger(__name__)

# A page grid is a 2D list of string values.
PageGrid = list[list[str]]

_MAX_COL_WIDTH = 60

# Every cell gets a thin grey outline so the table structure is visible.
_THIN_SIDE = Side(style="thin", color="9CA3AF")
_CELL_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)
_HEADER_FILL = PatternFill("solid", fgColor="E5E7EB")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(vertical="top", wrap_text=True)


def _autosize_columns(worksheet: Worksheet, grid: PageGrid) -> None:
    """Set column widths based on the longest value in each column."""
    if not grid:
        return
    n_cols = max((len(row) for row in grid), default=0)
    for col in range(n_cols):
        longest = 0
        for row in grid:
            if col < len(row):
                longest = max(longest, len(str(row[col])))
        letter = get_column_letter(col + 1)
        worksheet.column_dimensions[letter].width = min(longest + 2, _MAX_COL_WIDTH)


def _write_grid(worksheet: Worksheet, grid: PageGrid, bold_header: bool) -> None:
    """Write a 2D grid into a worksheet with visible cell borders.

    Every cell in the grid's rectangular extent is outlined — including cells
    that are empty — so the table structure stays readable in Excel.
    """
    if not grid:
        return
    n_cols = max(len(row) for row in grid)

    for r, row in enumerate(grid, start=1):
        for c in range(1, n_cols + 1):
            value = row[c - 1] if c - 1 < len(row) else ""
            cell = worksheet.cell(row=r, column=c, value=value)
            cell.border = _CELL_BORDER
            if bold_header and r == 1:
                cell.font = Font(bold=True)
                cell.fill = _HEADER_FILL
                cell.alignment = _HEADER_ALIGN
            else:
                cell.alignment = _CELL_ALIGN

    # Freeze the header row so it stays visible while scrolling.
    if bold_header:
        worksheet.freeze_panes = "A2"

    _autosize_columns(worksheet, grid)


def export_xlsx(
    pages: list[PageGrid],
    output_path: Path,
    merge: bool = False,
    bold_header: bool = True,
) -> Path:
    """Generate an ``.xlsx`` file from structured page grids.

    Args:
        pages: One 2D grid of string values per page.
        output_path: Destination ``.xlsx`` path.
        merge: If True, all pages are written into a single sheet; otherwise one
            sheet per page.
        bold_header: Bold the first row of each grid.

    Returns:
        The output path.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)  # drop the default empty sheet

    if merge:
        worksheet = workbook.create_sheet(title="Extraction")
        combined: PageGrid = []
        for grid in pages:
            combined.extend(grid)
        _write_grid(worksheet, combined, bold_header)
    else:
        for index, grid in enumerate(pages, start=1):
            worksheet = workbook.create_sheet(title=f"Page {index}")
            _write_grid(worksheet, grid, bold_header)

    if not workbook.sheetnames:  # guarantee at least one sheet
        workbook.create_sheet(title="Empty")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    log.info("excel.export", path=str(output_path), pages=len(pages), merge=merge)
    return output_path


def export_csv(pages: list[PageGrid], delimiter: str = ",") -> bytes:
    """Generate CSV bytes (UTF-8 with BOM) from structured page grids.

    All pages are concatenated; a blank line separates consecutive pages.

    Args:
        pages: One 2D grid of string values per page.
        delimiter: Field delimiter (``","`` or ``";"``).

    Returns:
        UTF-8-BOM-encoded CSV content as bytes.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=delimiter, lineterminator="\n")
    for i, grid in enumerate(pages):
        if i > 0:
            writer.writerow([])
        for row in grid:
            writer.writerow(row)
    return buffer.getvalue().encode("utf-8-sig")
