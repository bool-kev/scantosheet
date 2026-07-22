"""Unit tests for services that run without Tesseract/Poppler installed."""

from __future__ import annotations

from pathlib import Path

import cv2
import numpy as np
import openpyxl

from app.services import excel, pdf, table
from app.services.ocr import OcrResult, Word


def test_is_valid_pdf() -> None:
    assert pdf.is_valid_pdf(b"%PDF-1.7\n...")
    assert not pdf.is_valid_pdf(b"PK\x03\x04")  # a zip/xlsx, not a pdf
    assert not pdf.is_valid_pdf(b"")


def _word(text: str, left: int, line: int, conf: float = 90.0) -> Word:
    return Word(
        text=text,
        confidence=conf,
        left=left,
        top=line * 30,
        width=len(text) * 10,
        height=20,
        line_num=line,
        block_num=1,
        par_num=1,
    )


def test_structure_by_lines_fallback(tmp_path: Path) -> None:
    """With no grid image, structuring falls back to line-by-line splitting."""
    words = [
        _word("Name", left=0, line=1),
        _word("Age", left=300, line=1),
        _word("Alice", left=0, line=2),
        _word("30", left=300, line=2),
    ]
    result = OcrResult(text="Name Age\nAlice 30", words=words, mean_confidence=90.0)
    # Point at a non-existent image so grid detection yields nothing.
    grid = table.structure_page(tmp_path / "missing.png", result, tables_only=False)

    assert len(grid) == 2
    assert grid[0][0][0] == "Name"
    assert grid[0][1][0] == "Age"
    assert grid[1][0][0] == "Alice"
    assert grid[1][1][0] == "30"


def test_tables_only_ignores_free_text(tmp_path: Path) -> None:
    """With tables_only (default), a page without a table yields nothing."""
    words = [_word("Some", left=0, line=1), _word("paragraph", left=200, line=1)]
    result = OcrResult(text="Some paragraph", words=words, mean_confidence=90.0)
    assert table.structure_page(tmp_path / "missing.png", result) == []


def _draw_table_page(path: Path) -> tuple[int, int, int, int]:
    """Draw a page with a title, a 3x3 bordered table and a footer.

    Returns the table's (x0, y0, cell_width, cell_height) geometry.
    """
    h, w = 1200, 900
    img = np.full((h, w), 255, dtype=np.uint8)
    x0, y0, cw, ch = 100, 400, 220, 90
    for r in range(4):  # 4 horizontal rules -> 3 rows
        y = y0 + r * ch
        cv2.line(img, (x0, y), (x0 + 3 * cw, y), 0, 3)
    for c in range(4):  # 4 vertical rules -> 3 columns
        x = x0 + c * cw
        cv2.line(img, (x, y0), (x, y0 + 3 * ch), 0, 3)
    cv2.imwrite(str(path), img)
    return x0, y0, cw, ch


def _word_at(text: str, cx: int, cy: int, conf: float = 90.0) -> Word:
    """Build a Word centred on (cx, cy)."""
    width, height = len(text) * 8, 20
    return Word(
        text=text,
        confidence=conf,
        left=cx - width // 2,
        top=cy - height // 2,
        width=width,
        height=height,
        line_num=1,
        block_num=1,
        par_num=1,
    )


def test_only_words_inside_table_are_kept(tmp_path: Path) -> None:
    """Text above and below the table is discarded; in-table cells are mapped."""
    image = tmp_path / "page.png"
    x0, y0, cw, ch = _draw_table_page(image)

    def centre(row: int, col: int) -> tuple[int, int]:
        return x0 + col * cw + cw // 2, y0 + row * ch + ch // 2

    words = [
        # Outside the table: a title above and a footer below.
        _word_at("MINISTERE", 450, 100),
        _word_at("BURKINA", 450, 250),
        _word_at("SignatureFooter", 450, 1100),
        # Inside the table.
        _word_at("Produit", *centre(0, 0)),
        _word_at("Qte", *centre(0, 1)),
        _word_at("Prix", *centre(0, 2)),
        _word_at("Pommes", *centre(1, 0)),
        _word_at("12", *centre(1, 1)),
        _word_at("3.50", *centre(1, 2)),
    ]
    result = OcrResult(text="", words=words, mean_confidence=90.0)
    grid = table.structure_page(image, result)

    flat = [cell[0] for row in grid for cell in row]
    assert "MINISTERE" not in flat
    assert "BURKINA" not in flat
    assert "SignatureFooter" not in flat
    assert grid[0][0][0] == "Produit"
    assert grid[0][2][0] == "Prix"
    assert grid[1][0][0] == "Pommes"
    assert grid[1][2][0] == "3.50"


def test_trim_empty_drops_blank_rows_and_columns() -> None:
    grid = [
        [("A", 90.0), ("", 0.0), ("B", 80.0)],
        [("", 0.0), ("", 0.0), ("", 0.0)],
        [("C", 70.0), ("", 0.0), ("D", 60.0)],
    ]
    trimmed = table._trim_empty(grid)
    assert trimmed == [[("A", 90.0), ("B", 80.0)], [("C", 70.0), ("D", 60.0)]]


def test_export_xlsx_one_sheet_per_page(tmp_path: Path) -> None:
    pages = [
        [["A", "B"], ["1", "2"]],
        [["C", "D"], ["3", "4"]],
    ]
    out = tmp_path / "out.xlsx"
    excel.export_xlsx(pages, out, merge=False)

    workbook = openpyxl.load_workbook(out)
    assert workbook.sheetnames == ["Page 1", "Page 2"]
    assert workbook["Page 1"]["A1"].value == "A"
    assert workbook["Page 1"]["A1"].font.bold
    assert workbook["Page 2"]["B2"].value == "4"


def test_export_xlsx_draws_cell_borders(tmp_path: Path) -> None:
    """Every cell in the grid extent is outlined, including short/empty rows."""
    pages = [[["Produit", "Qte", "Prix"], ["Pommes"]]]  # 2nd row is ragged
    out = tmp_path / "borders.xlsx"
    excel.export_xlsx(pages, out)

    ws = openpyxl.load_workbook(out)["Page 1"]
    for coord in ("A1", "B1", "C1", "A2", "B2", "C2"):
        border = ws[coord].border
        assert border.left.style == "thin", f"{coord} missing left border"
        assert border.right.style == "thin", f"{coord} missing right border"
        assert border.top.style == "thin", f"{coord} missing top border"
        assert border.bottom.style == "thin", f"{coord} missing bottom border"

    # Header styling and frozen header row.
    assert ws["A1"].font.bold
    assert ws["A1"].alignment.wrap_text
    assert ws.freeze_panes == "A2"
    # Ragged row was padded with an empty cell that still carries a border
    # (openpyxl reads a written empty string back as None).
    assert ws["B2"].value in (None, "")
    assert ws.max_column == 3


def test_export_xlsx_merged(tmp_path: Path) -> None:
    pages = [[["A"]], [["B"]]]
    out = tmp_path / "merged.xlsx"
    excel.export_xlsx(pages, out, merge=True)

    workbook = openpyxl.load_workbook(out)
    assert workbook.sheetnames == ["Extraction"]
    assert workbook["Extraction"]["A1"].value == "A"
    assert workbook["Extraction"]["A2"].value == "B"


def test_export_csv_bom_and_delimiter() -> None:
    pages = [[["a", "b"], ["c", "d"]]]
    content = excel.export_csv(pages, delimiter=";")
    assert content.startswith(b"\xef\xbb\xbf")  # UTF-8 BOM
    text = content.decode("utf-8-sig")
    assert "a;b" in text
    assert "c;d" in text
