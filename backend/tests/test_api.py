"""API smoke tests using FastAPI's TestClient.

These cover the request/response wiring that does not require Tesseract or
Poppler (health, listing, validation errors, 404s). The full OCR path is
exercised via Docker.
"""

from __future__ import annotations

from collections.abc import Iterator

import pytest
from fastapi.testclient import TestClient

from app.config import get_settings
from app.main import app


@pytest.fixture(scope="module")
def client() -> Iterator[TestClient]:
    """These tests exercise business logic, not auth — run with auth disabled.

    Auth itself (open vs. enabled, scoping, admin) is covered exhaustively in
    test_auth.py. Mutated directly (not via ``monkeypatch``) because this
    fixture is module-scoped and ``monkeypatch`` is function-scoped.
    """
    settings = get_settings()
    original = settings.auth_enabled
    settings.auth_enabled = False
    try:
        # Entering the context manager triggers the lifespan (init_db).
        with TestClient(app) as test_client:
            yield test_client
    finally:
        settings.auth_enabled = original


def test_health(client: TestClient) -> None:
    response = client.get("/api/health")
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "ok"
    assert "tesseract_available" in body


def test_list_documents_empty_ok(client: TestClient) -> None:
    response = client.get("/api/documents")
    assert response.status_code == 200
    body = response.json()
    assert set(body) >= {"items", "total", "page", "page_size"}


def test_upload_rejects_non_pdf(client: TestClient) -> None:
    response = client.post(
        "/api/documents",
        files={"file": ("note.txt", b"just some text", "text/plain")},
        data={"language": "fra", "preprocessing": "true"},
    )
    assert response.status_code == 400
    assert "not a valid PDF" in response.json()["detail"]


def test_upload_rejects_empty_file(client: TestClient) -> None:
    response = client.post(
        "/api/documents",
        files={"file": ("empty.pdf", b"", "application/pdf")},
        data={"language": "fra"},
    )
    assert response.status_code == 400


def test_upload_rejects_invalid_callback_url(client: TestClient) -> None:
    response = client.post(
        "/api/documents",
        files={"file": ("scan.pdf", b"%PDF-1.7\nfake", "application/pdf")},
        data={"language": "fra", "callback_url": "not-a-url"},
    )
    assert response.status_code == 400
    assert "callback_url" in response.json()["detail"]


def test_download_uses_merge_pages_from_upload(client: TestClient) -> None:
    """merge_pages chosen at import drives the export layout, and can be overridden."""
    import openpyxl

    from app.database import SessionLocal
    from app.models import Document, DocumentStatus, Page

    db = SessionLocal()
    doc = Document(
        filename="merged.pdf",
        stored_filename="merged.pdf",
        status=DocumentStatus.DONE,
        page_count=2,
        language="fra",
        preprocessing=True,
        merge_pages=True,
    )
    for number, value in ((1, "A"), (2, "B")):
        doc.pages.append(
            Page(
                page_number=number,
                text=value,
                data_json=f'[[{{"value": "{value}", "confidence": 90.0}}]]',
                mean_confidence=90.0,
            )
        )
    db.add(doc)
    db.commit()
    doc_id = doc.id
    db.close()

    assert client.get(f"/api/documents/{doc_id}").json()["merge_pages"] is True

    # No query param -> the document's own setting wins: a single sheet.
    response = client.get(f"/api/documents/{doc_id}/download")
    assert response.status_code == 200
    from io import BytesIO

    merged = openpyxl.load_workbook(BytesIO(response.content))
    assert merged.sheetnames == ["Extraction"]
    assert [c[0].value for c in merged["Extraction"].iter_rows()] == ["A", "B"]

    # Explicit override still wins.
    split = openpyxl.load_workbook(
        BytesIO(client.get(f"/api/documents/{doc_id}/download?merge=false").content)
    )
    assert split.sheetnames == ["Page 1", "Page 2"]


def test_cors_allows_configured_origin(client: TestClient) -> None:
    """A browser client from an allowed origin must get CORS headers back."""
    origin = "http://localhost:5173"
    response = client.options(
        "/api/documents",
        headers={
            "Origin": origin,
            "Access-Control-Request-Method": "POST",
            "Access-Control-Request-Headers": "x-api-key",
        },
    )
    assert response.status_code == 200
    assert response.headers["access-control-allow-origin"] == origin
    assert "x-api-key" in response.headers["access-control-allow-headers"].lower()


def test_cors_exposes_content_disposition(client: TestClient) -> None:
    """Cross-origin JS cannot read the download filename without this header."""
    response = client.get("/api/documents", headers={"Origin": "http://localhost:5173"})
    assert response.status_code == 200
    assert "content-disposition" in response.headers["access-control-expose-headers"].lower()


def test_get_missing_document_404(client: TestClient) -> None:
    response = client.get("/api/documents/999999")
    assert response.status_code == 404


def test_get_document_with_pages_serializes(client: TestClient) -> None:
    """Regression: GET /documents/{id} must serialize ORM pages into PageData."""
    from app.database import SessionLocal
    from app.models import Document, DocumentStatus, Page

    db = SessionLocal()
    doc = Document(
        filename="serialize.pdf",
        stored_filename="serialize.pdf",
        status=DocumentStatus.DONE,
        page_count=1,
        language="fra",
        preprocessing=True,
    )
    doc.pages.append(
        Page(
            page_number=1,
            text="Produit Prix",
            data_json='[[{"value": "Produit", "confidence": 95.0}, '
            '{"value": "Prix", "confidence": 88.0}]]',
            mean_confidence=91.5,
        )
    )
    db.add(doc)
    db.commit()
    doc_id = doc.id
    db.close()

    response = client.get(f"/api/documents/{doc_id}")
    assert response.status_code == 200
    body = response.json()
    assert body["page_count"] == 1
    assert body["pages"][0]["data"][0][0]["value"] == "Produit"
    assert body["pages"][0]["data"][0][0]["confidence"] == 95.0

    # And the preview endpoint returns the same structured data.
    preview = client.get(f"/api/documents/{doc_id}/preview")
    assert preview.status_code == 200
    assert preview.json()["pages"][0]["data"][0][1]["value"] == "Prix"
